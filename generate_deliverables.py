"""
Generate submission deliverables for all 10 papers.

Primary output per paper:
  {AuthorYear}_Report.xlsx  — 19-sheet workbook (source + 00_PaperInfo + IterationDetail)
  Paper_Info_Record.pdf     — Appendix A template (6 sections)

Both saved inside paper_dir/ and copied to repo root.

Usage:
  python generate_deliverables.py              # process all papers
  python generate_deliverables.py --paper 0005 # process one paper
"""

from __future__ import annotations

import argparse
import re
import shutil
from datetime import date
from itertools import product
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    )
    _REPORTLAB_OK = True
except ImportError:
    _REPORTLAB_OK = False

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
REPO_ROOT = Path(__file__).parent
PAPER_OUTPUT = REPO_ROOT / "paper_analysis_output"

# ---------------------------------------------------------------------------
# Paper configuration
# ---------------------------------------------------------------------------
PAPERS = {
    "0005": {
        "paper_dir":          PAPER_OUTPUT / "Stroube2025",
        "workbook":           PAPER_OUTPUT / "Stroube2025" / "full_run" / "Stroube2025_Report.xlsx",
        "paper_info":         PAPER_OUTPUT / "Stroube2025" / "paper_info.xlsx",
        "author_year":        "Stroube2025",
        "focal_iv":           "log_pop_black_aa",
        "baseline_coef":      0.0307,
        "baseline_se":        0.0054,
        "baseline_pval":      2.80e-08,
        "regression_outputs": PAPER_OUTPUT / "Stroube2025" / "regression_outputs",
        "key_vars": [
            "log_pop_black_aa",
            "log_total_bachelor_deg",
            "log_pop_total_poverty",
            "log_total_social_cap",
        ],
        "mechanisms":   ["MCAR", "MAR", "NMAR"],
        "pct_strings":  ["1pct", "5pct", "10pct", "20pct", "30pct", "40pct", "50pct"],
        "methods":      ["LD", "Mean", "Reg", "Iter", "RF", "DL", "MILGBM"],
        "n_iters":      30,
        "txt_has_data": True,   # re-run (2026-04) produced real pyfixest tidy() outputs
    },
    "0017": {
        "paper_dir":          PAPER_OUTPUT / "Stroube2024",
        "workbook":           PAPER_OUTPUT / "Stroube2024" / "full_run" / "Stroube2024_Report.xlsx",
        "paper_info":         PAPER_OUTPUT / "Stroube2024" / "paper_info.xlsx",
        "author_year":        "Stroube2024",
        "focal_iv":           "FLead",
        "baseline_coef":      0.0468,
        "baseline_se":        0.0080,
        "baseline_pval":      6.23e-09,
        "regression_outputs": PAPER_OUTPUT / "Stroube2024" / "regression_outputs",
        "key_vars": [
            "kim_violence_gore",
            "kim_sex_nudity",
            "kim_language",
            "log_bom_opening_theaters",
        ],
        "mechanisms":   ["MCAR", "MAR", "NMAR"],
        "pct_strings":  ["1pct", "5pct", "10pct", "20pct", "30pct", "40pct", "50pct"],
        "methods":      ["LD", "Mean", "Reg", "Iter", "RF", "DL", "MILGBM"],
        "n_iters":      30,
        "txt_has_data": True,   # .txt files have full OLS summaries
    },
    "0019": {
        "paper_dir":          PAPER_OUTPUT / "Greene2021",
        "workbook":           PAPER_OUTPUT / "Greene2021" / "full_run" / "Greene2021_Report.xlsx",
        "paper_info":         PAPER_OUTPUT / "Greene2021" / "paper_info.xlsx",
        "author_year":        "Greene2021",
        "focal_iv":           "ad_law2",
        "baseline_coef":      -0.0110,
        "baseline_se":        0.0034,
        "baseline_pval":      0.0013,
        "regression_outputs": PAPER_OUTPUT / "Greene2021" / "regression_outputs",
        "key_vars":           ["ln_at_adj", "ppent_at", "state_inc_growth"],
        "mechanisms":   ["MCAR", "MAR", "NMAR"],
        "pct_strings":  ["1pct", "5pct", "10pct", "20pct", "30pct", "40pct", "50pct"],
        "methods":      ["LD", "Mean", "Reg", "Iter", "RF", "DL", "MILGBM"],
        "n_iters":      30,
        "txt_has_data": True,
    },
    "0020": {
        "paper_dir":          PAPER_OUTPUT / "Meyer2024",
        "workbook":           PAPER_OUTPUT / "Meyer2024" / "full_run" / "Meyer2024_Report.xlsx",
        "paper_info":         PAPER_OUTPUT / "Meyer2024" / "paper_info.xlsx",
        "author_year":        "Meyer2024",
        "focal_iv":           "afterXVGM",
        "baseline_coef":      1.1369,
        "baseline_se":        0.4846,
        "baseline_pval":      0.020,
        "regression_outputs": PAPER_OUTPUT / "Meyer2024" / "regression_outputs",
        "key_vars":           ["post_lav_Visits", "post_lHerfCont", "post_lav_Visits_VGM"],
        "mechanisms":   ["MCAR", "MAR", "NMAR"],
        "pct_strings":  ["1pct", "5pct", "10pct", "20pct", "30pct", "40pct", "50pct"],
        "methods":      ["LD", "Mean", "Reg", "Iter", "RF", "DL", "MILGBM"],
        "n_iters":      30,
        "txt_has_data": True,
    },
    "0022": {
        "paper_dir":          PAPER_OUTPUT / "Santamaria2024",
        "workbook":           PAPER_OUTPUT / "Santamaria2024" / "full_run" / "Santamaria2024_Report.xlsx",
        "paper_info":         PAPER_OUTPUT / "Santamaria2024" / "paper_info.xlsx",
        "author_year":        "Santamaria2024",
        "focal_iv":           "Post_x_Treatment",
        "baseline_coef":      0.3739,
        "baseline_se":        0.1913,
        "baseline_pval":      0.052,
        "regression_outputs": PAPER_OUTPUT / "Santamaria2024" / "regression_outputs",
        "key_vars":           ["Age", "WorkExp", "EntrepExperience"],
        "mechanisms":   ["MCAR", "MAR", "NMAR"],
        "pct_strings":  ["1pct", "5pct", "10pct", "20pct", "30pct", "40pct", "50pct"],
        "methods":      ["LD", "Mean", "Reg", "Iter", "RF", "DL", "MILGBM"],
        "n_iters":      30,
        "txt_has_data": True,
    },
    "0023": {
        "paper_dir":          PAPER_OUTPUT / "Chyz2023",
        "workbook":           PAPER_OUTPUT / "Chyz2023" / "full_run" / "Chyz2023_Report.xlsx",
        "paper_info":         PAPER_OUTPUT / "Chyz2023" / "paper_info.xlsx",
        "author_year":        "Chyz2023",
        "focal_iv":           "diff_laggaap_etr",
        "baseline_coef":      -0.0409,
        "baseline_se":        0.00876,
        "baseline_pval":      3.03e-06,
        "regression_outputs": PAPER_OUTPUT / "Chyz2023" / "regression_outputs",
        "key_vars":           ["ch1_s_rd", "ch1_s_sga", "ch1_roa_pretax", "ch1_size"],
        "mechanisms":   ["MCAR", "MAR", "NMAR"],
        "pct_strings":  ["1pct", "5pct", "10pct", "20pct", "30pct", "40pct", "50pct"],
        "methods":      ["LD", "Mean", "Reg", "Iter", "RF", "DL", "MILGBM"],
        "n_iters":      30,
        "txt_has_data": True,
        # Source workbook Coef_Stability_Summary and Mean_Stability sheets are incomplete
        # because the run was interrupted and resumed — in-memory results only cover the
        # restart session. Rebuild those sheets from the parsed IterationDetail data.
        "rebuild_summary": True,
    },
    "0025": {
        "paper_dir":          PAPER_OUTPUT / "Anderson2018",
        "workbook":           PAPER_OUTPUT / "Anderson2018" / "full_run" / "Anderson2018_Report.xlsx",
        "paper_info":         PAPER_OUTPUT / "Anderson2018" / "paper_info.xlsx",
        "author_year":        "Anderson2018",
        "focal_iv":           "Treatment_FIN",
        "baseline_coef":      2230.74,
        "baseline_se":        1052.65,
        "baseline_pval":      0.0345,
        "regression_outputs": PAPER_OUTPUT / "Anderson2018" / "regression_outputs",
        "key_vars":           ["Age", "Operating_yearstotal", "Activity_Hours", "pre_Profits3_composite_w1"],
        "mechanisms":   ["MCAR", "MAR", "NMAR"],
        "pct_strings":  ["1pct", "5pct", "10pct", "20pct", "30pct", "40pct", "50pct"],
        "methods":      ["LD", "Mean", "Reg", "Iter", "RF", "DL", "MILGBM"],
        "n_iters":      30,
        "txt_has_data": True,
        # Source workbook has empty summary sheets — final restart skipped all combos
        # via done_set (all 17,640 unique keys present after parallel-run duplicates).
        # Rebuild from the complete txt files.
        "rebuild_summary": True,
    },
    "0018": {
        "paper_dir":          PAPER_OUTPUT / "Fang2022",
        "workbook":           PAPER_OUTPUT / "Fang2022" / "full_run" / "Fang2022_Report.xlsx",
        "paper_info":         PAPER_OUTPUT / "Fang2022" / "paper_info.xlsx",
        "author_year":        "Fang2022",
        "focal_iv":           "lrdefficiency_postremoval",
        "baseline_coef":      0.00595,
        "baseline_se":        0.00357,
        "baseline_pval":      0.0959,
        "regression_outputs": PAPER_OUTPUT / "Fang2022" / "regression_outputs",
        "key_vars":           ["lroa", "ltobinq", "lleverage"],
        "mechanisms":   ["MCAR", "MAR", "NMAR"],
        "pct_strings":  ["1pct", "5pct", "10pct", "20pct", "30pct", "40pct", "50pct"],
        "methods":      ["LD", "Mean", "Reg", "Iter", "RF", "DL", "MILGBM"],
        "n_iters":      30,
        "txt_has_data": True,
        # Source workbook is complete (single uninterrupted session) — no rebuild needed.
    },
    "0024": {
        "paper_dir":          PAPER_OUTPUT / "Christensen2021",
        "workbook":           PAPER_OUTPUT / "Christensen2021" / "full_run" / "Christensen2021_Report.xlsx",
        "paper_info":         PAPER_OUTPUT / "Christensen2021" / "paper_info.xlsx",
        "author_year":        "Christensen2021",
        "focal_iv":           "politicalhedge",
        "baseline_coef":      -0.0311,
        "baseline_se":        0.00938,
        "baseline_pval":      0.000933,
        "regression_outputs": PAPER_OUTPUT / "Christensen2021" / "regression_outputs",
        "key_vars":           ["mktvol", "beta", "btm", "competition"],
        "mechanisms":   ["MCAR", "MAR", "NMAR"],
        "pct_strings":  ["1pct", "5pct", "10pct", "20pct", "30pct", "40pct", "50pct"],
        "methods":      ["LD", "Mean", "Reg", "Iter", "RF", "DL", "MILGBM"],
        "n_iters":      30,
        "txt_has_data": True,
        # Source workbook complete (single uninterrupted session) — no rebuild needed.
    },
    "0021": {
        "paper_dir":          PAPER_OUTPUT / "Hu2025",
        "workbook":           PAPER_OUTPUT / "Hu2025" / "full_run" / "Hu2025_Report.xlsx",
        "paper_info":         PAPER_OUTPUT / "Hu2025" / "paper_info.xlsx",
        "author_year":        "Hu2025",
        "focal_iv":           "post1_x_treat1",
        "baseline_coef":      0.0331,
        "baseline_se":        0.00457,
        "baseline_pval":      5.78e-13,
        "regression_outputs": PAPER_OUTPUT / "Hu2025" / "regression_outputs",
        "key_vars":           ["ptita", "mb", "lev", "aret"],
        "mechanisms":   ["MCAR", "MAR", "NMAR"],
        "pct_strings":  ["1pct", "5pct", "10pct", "20pct", "30pct", "40pct", "50pct"],
        "methods":      ["LD", "Mean", "Reg", "Iter", "RF", "DL", "MILGBM"],
        "n_iters":      30,
        "txt_has_data": True,
        # Run was interrupted and resumed — in-memory workbook only covers the last session.
        # Rebuild summary sheets from the complete txt files.
        "rebuild_summary": True,
    },
}

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
HEADER_FONT  = Font(bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="1E50A0")
LABEL_FONT   = Font(bold=True)
LABEL_FILL   = PatternFill("solid", fgColor="F0F0F8")
ALT_ROW_FILL = PatternFill("solid", fgColor="F5F5F5")


def _style_header_row(ws, row_idx: int = 1) -> None:
    for cell in ws[row_idx]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=False, vertical="center")


# ---------------------------------------------------------------------------
# Fix A — read 00_PaperInfo rows from paper_info_XXXX.xlsx
# ---------------------------------------------------------------------------
def _load_paper_info_rows(paper_info_path: Path) -> list[tuple[str, str]]:
    df = pd.read_excel(paper_info_path, sheet_name=0)
    rows = []
    for _, row in df.iterrows():
        field = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
        value = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ""
        if field and field != "Field":
            rows.append((field, value))
    return rows


def _build_paper_info_sheet(wb: openpyxl.Workbook, rows: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet("00_PaperInfo")

    ws.append(["Field", "Value"])
    _style_header_row(ws)

    for field, value in rows:
        ws.append([field, value])
        r = ws.max_row
        ws[f"A{r}"].font  = LABEL_FONT
        ws[f"A{r}"].fill  = LABEL_FILL
        ws[f"A{r}"].alignment = Alignment(wrap_text=True)
        ws[f"B{r}"].alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80
    ws.freeze_panes = "A2"

    # Move to position 0 (first tab)
    wb.move_sheet("00_PaperInfo", offset=-len(wb.sheetnames) + 1)


# ---------------------------------------------------------------------------
# Fix B — IterationDetail sheet
# ---------------------------------------------------------------------------

def _parse_ols_txt(path: Path, focal_iv: str) -> dict | None:
    """
    Parse a regression output .txt file.
    Handles two formats:
      1. statsmodels OLS summary table (FLead row in coefficients block)
      2. MI Pooled Result format (Pooled Coef: / Pooled SE: / Pooled pval: / N (mean):)
    Returns dict with coef, se, pval, nobs or None on failure.
    """
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return None

    if not text or text.strip() in ("None", ""):
        return None

    # --- Try MI Pooled format first ---
    if "Pooled Coef:" in text:
        try:
            coef = float(re.search(r'Pooled Coef:\s*([-\d.e+]+)', text).group(1))
            se   = float(re.search(r'Pooled SE:\s*([-\d.e+]+)', text).group(1))
            pval = float(re.search(r'Pooled pval:\s*([-\d.e+]+)', text).group(1))
            m_n  = re.search(r'N \(mean\):\s*([\d.]+)', text)
            nobs = int(float(m_n.group(1))) if m_n else None
            return {"coef": coef, "se": se, "pval": pval, "nobs": nobs}
        except (AttributeError, ValueError):
            pass

    # --- Try custom pyfixest header format (Paper 0005 re-run) ---
    # Format: lines like "Coef: 0.029300\nSE: 0.004601\npval: 0.000000\nNobs: 32647"
    if re.search(r'^Coef:\s', text, re.MULTILINE):
        try:
            coef = float(re.search(r'^Coef:\s*([-\d.e+]+)', text, re.MULTILINE).group(1))
            se   = float(re.search(r'^SE:\s*([-\d.e+]+)', text, re.MULTILINE).group(1))
            pval = float(re.search(r'^pval:\s*([-\d.e+]+)', text, re.MULTILINE).group(1))
            m_n  = re.search(r'^Nobs:\s*([\d]+)', text, re.MULTILINE)
            nobs = int(m_n.group(1)) if m_n else None
            return {"coef": coef, "se": se, "pval": pval, "nobs": nobs}
        except (AttributeError, ValueError):
            pass

    # --- Try OLS summary table format ---
    # Extract nobs
    nobs = None
    m = re.search(r'No\.\s+Observations:\s+([\d,]+)', text)
    if m:
        nobs = int(m.group(1).replace(",", ""))

    # Find focal IV row in coefficient table
    # Format: "FLead                        0.0477      0.008      5.934      0.000 ..."
    pattern = re.compile(
        r'^\s*' + re.escape(focal_iv) + r'\s+([-\d.e+]+)\s+([-\d.e+]+)\s+([-\d.e+]+)\s+([-\d.e+]+)',
        re.MULTILINE
    )
    m = pattern.search(text)
    if not m:
        return None

    try:
        coef = float(m.group(1))
        se   = float(m.group(2))
        # group(3) is t-stat; group(4) is P>|t|
        pval = float(m.group(4))
    except (ValueError, IndexError):
        return None

    return {"coef": coef, "se": se, "pval": pval, "nobs": nobs}


def _build_iteration_detail_0017(cfg: dict) -> pd.DataFrame:
    """
    Build IterationDetail for paper 0017 by parsing .txt files.
    """
    reg_dir = cfg["regression_outputs"]
    focal_iv = cfg["focal_iv"]
    baseline_coef = cfg["baseline_coef"]
    baseline_se   = cfg["baseline_se"]
    baseline_pval = cfg["baseline_pval"]

    records = []
    total_expected = (
        len(cfg["key_vars"]) * len(cfg["mechanisms"]) *
        len(cfg["pct_strings"]) * len(cfg["methods"]) * cfg["n_iters"]
    )
    print(f"    Parsing .txt files (expected {total_expected:,} files)...")

    parsed_ok = 0
    parsed_fail = 0

    for mechanism, pct_str, key_var, method, iteration in product(
        cfg["mechanisms"], cfg["pct_strings"], cfg["key_vars"],
        cfg["methods"], range(cfg["n_iters"])
    ):
        txt_path = reg_dir / mechanism / pct_str / key_var / method / f"iter{iteration}_model_{key_var}.txt"
        parsed = _parse_ols_txt(txt_path, focal_iv)

        if parsed:
            coef = parsed["coef"]
            se   = parsed["se"]
            pval = parsed["pval"]
            nobs = parsed["nobs"]
            sign = int(np.sign(coef)) if not np.isnan(coef) else np.nan
            significant = int(pval < 0.05) if not np.isnan(pval) else np.nan
            coef_delta  = coef - baseline_coef
            sign_match  = int(np.sign(coef) == np.sign(baseline_coef))
            sig_match   = int((pval < 0.05) == (baseline_pval < 0.05))
            both_match  = int(bool(sign_match) and bool(sig_match))
            parsed_ok += 1
        else:
            coef = se = pval = coef_delta = np.nan
            nobs = np.nan
            sign = significant = sign_match = sig_match = both_match = np.nan
            parsed_fail += 1

        records.append({
            "key_var":       key_var,
            "mechanism":     mechanism,
            "pct_str":       pct_str,
            "method":        method,
            "iteration":     iteration,
            "coef":          coef,
            "se":            se,
            "pval":          pval,
            "nobs":          nobs,
            "sign":          sign,
            "significant":   significant,
            "baseline_coef": baseline_coef,
            "baseline_se":   baseline_se,
            "baseline_pval": baseline_pval,
            "coef_delta":    coef_delta,
            "sign_match":    sign_match,
            "sig_match":     sig_match,
            "both_match":    both_match,
        })

    print(f"    Parsed OK: {parsed_ok:,} | Failed/empty: {parsed_fail:,}")
    df = pd.DataFrame(records)
    df = df.sort_values(["mechanism", "pct_str", "key_var", "method", "iteration"]).reset_index(drop=True)
    return df


def _build_iteration_detail_0005(cfg: dict) -> pd.DataFrame:
    """
    Build IterationDetail skeleton for paper 0005.
    Coefficient data was not retained in the simulation outputs (.txt files are empty).
    Structural columns are populated; coef/se/pval are NaN.
    """
    baseline_coef = cfg["baseline_coef"]
    baseline_se   = cfg["baseline_se"]
    baseline_pval = cfg["baseline_pval"]

    records = []
    for mechanism, pct_str, key_var, method, iteration in product(
        cfg["mechanisms"], cfg["pct_strings"], cfg["key_vars"],
        cfg["methods"], range(cfg["n_iters"])
    ):
        records.append({
            "key_var":       key_var,
            "mechanism":     mechanism,
            "pct_str":       pct_str,
            "method":        method,
            "iteration":     iteration,
            "coef":          np.nan,
            "se":            np.nan,
            "pval":          np.nan,
            "nobs":          np.nan,
            "sign":          np.nan,
            "significant":   np.nan,
            "baseline_coef": baseline_coef,
            "baseline_se":   baseline_se,
            "baseline_pval": baseline_pval,
            "coef_delta":    np.nan,
            "sign_match":    np.nan,
            "sig_match":     np.nan,
            "both_match":    np.nan,
            "note":          "Coefficient data not retained in simulation outputs (pyfixest .summary() returned None)",
        })

    df = pd.DataFrame(records)
    df = df.sort_values(["mechanism", "pct_str", "key_var", "method", "iteration"]).reset_index(drop=True)
    print(f"    Built structural skeleton: {len(df):,} rows (coef columns are NaN)")
    return df


def _write_iteration_detail_sheet(wb: openpyxl.Workbook, df: pd.DataFrame) -> None:
    """Append IterationDetail sheet to workbook with formatting."""
    ws = wb.create_sheet("IterationDetail")

    # Write header
    cols = list(df.columns)
    ws.append(cols)
    _style_header_row(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    # Write rows with alternate shading
    for i, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = ALT_ROW_FILL

    # Set column widths
    width_map = {
        "key_var": 26, "mechanism": 10, "pct_str": 8, "method": 8,
        "iteration": 9, "coef": 12, "se": 12, "pval": 12, "nobs": 8,
        "sign": 6, "significant": 11, "baseline_coef": 14, "baseline_se": 12,
        "baseline_pval": 14, "coef_delta": 12, "sign_match": 11,
        "sig_match": 10, "both_match": 11, "note": 55,
    }
    for col_idx, col_name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width_map.get(col_name, 12)

    print(f"    IterationDetail: {len(df):,} rows written")


# ---------------------------------------------------------------------------
# Fix C — rebuild summary sheets from IterationDetail (for resumed runs)
# ---------------------------------------------------------------------------

def _recompute_coef_stability_summary(df_iter: pd.DataFrame) -> pd.DataFrame:
    """Recompute Coef_Stability_Summary from IterationDetail df."""
    records = []
    for (kv, mech, pct, method), g in df_iter.groupby(
        ["key_var", "mechanism", "pct_str", "method"], sort=False
    ):
        valid = g.dropna(subset=["both_match"])
        n = len(valid)
        both_same = int(valid["both_match"].sum()) if n > 0 else 0
        ss_n = int(((valid["sign_match"] == 1) & (valid["sig_match"] == 0)).sum()) if n > 0 else 0

        if n > 0:
            b_prop = round(both_same / n * 100, 1)
            ss_prop = round(ss_n / n * 100, 1)
            # Wilson CI for B_prop
            p = both_same / n
            z = 1.96
            denom = 1 + z ** 2 / n
            center = (p + z ** 2 / (2 * n)) / denom
            margin = z * np.sqrt(p * (1 - p) / n + z ** 2 / (4 * n ** 2)) / denom
            b_ci_lo = round(max(0.0, center - margin) * 100, 1)
            b_ci_hi = round(min(100.0, center + margin) * 100, 1)
            mean_nobs = int(valid["nobs"].mean()) if valid["nobs"].notna().any() else None
        else:
            b_prop = ss_prop = b_ci_lo = b_ci_hi = mean_nobs = None

        records.append({
            "KeyVar": kv, "Mechanism": mech, "Proportion": pct, "Method": method,
            "N_iters": n, "BothSame": both_same, "SignSameSigChanged": ss_n,
            "B_prop": b_prop, "SS_prop": ss_prop,
            "B_CI_lo": b_ci_lo, "B_CI_hi": b_ci_hi,
            "Mean_N_obs": mean_nobs,
        })

    PCT_ORDER = ["1pct", "5pct", "10pct", "20pct", "30pct", "40pct", "50pct"]
    MECH_ORDER = ["MCAR", "MAR", "NMAR"]
    df = pd.DataFrame(records)
    df["_mech_ord"] = df["Mechanism"].map({m: i for i, m in enumerate(MECH_ORDER)})
    df["_pct_ord"] = df["Proportion"].map({p: i for i, p in enumerate(PCT_ORDER)})
    df = df.sort_values(["_mech_ord", "_pct_ord", "KeyVar", "Method"]).drop(
        columns=["_mech_ord", "_pct_ord"]
    ).reset_index(drop=True)
    return df


def _recompute_mean_stability(df_css: pd.DataFrame, mechanism: str) -> pd.DataFrame:
    """Recompute Mean_Stability_<MECH> from Coef_Stability_Summary df."""
    PCT_ORDER = ["1pct", "5pct", "10pct", "20pct", "30pct", "40pct", "50pct"]
    sub = df_css[df_css["Mechanism"] == mechanism].copy()

    rows = []
    for (kv, method), g in sub.groupby(["KeyVar", "Method"], sort=False):
        row = {"KeyVar": kv, "Method": method}
        for pct in PCT_ORDER:
            r = g[g["Proportion"] == pct]
            if len(r) == 1:
                row[f"B_{pct}"]    = r.iloc[0]["B_prop"]
                row[f"B_lo_{pct}"] = r.iloc[0]["B_CI_lo"]
                row[f"B_hi_{pct}"] = r.iloc[0]["B_CI_hi"]
                row[f"SS_{pct}"]   = r.iloc[0]["SS_prop"]
            else:
                row[f"B_{pct}"] = row[f"B_lo_{pct}"] = row[f"B_hi_{pct}"] = row[f"SS_{pct}"] = None
        rows.append(row)

    col_order = ["KeyVar", "Method"]
    for pct in PCT_ORDER:
        col_order += [f"B_{pct}", f"B_lo_{pct}", f"B_hi_{pct}", f"SS_{pct}"]
    df = pd.DataFrame(rows)
    df = df.sort_values(["KeyVar", "Method"]).reset_index(drop=True)
    return df[[c for c in col_order if c in df.columns]]


def _recompute_model_comparison(df_iter: pd.DataFrame) -> pd.DataFrame:
    """Recompute Model_Comparison (RMSE + avg_rel_se per method/mechanism/pct)."""
    records = []
    for (method, mech, pct), g in df_iter.groupby(["method", "mechanism", "pct_str"], sort=False):
        valid = g.dropna(subset=["coef", "se"])
        if len(valid) == 0:
            continue
        rmse = float(np.sqrt(((valid["coef_delta"]) ** 2).mean()))
        baseline_se = valid["baseline_se"].iloc[0]
        avg_rel_se = float((valid["se"] / baseline_se).mean()) if baseline_se else None
        records.append({
            "method": method, "mechanism": mech, "pct_str": pct,
            "rmse": rmse, "avg_rel_se": avg_rel_se, "n_iters": len(valid),
        })
    PCT_ORDER = ["1pct", "5pct", "10pct", "20pct", "30pct", "40pct", "50pct"]
    MECH_ORDER = ["MCAR", "MAR", "NMAR"]
    df = pd.DataFrame(records)
    df["_mord"] = df["mechanism"].map({m: i for i, m in enumerate(MECH_ORDER)})
    df["_pord"] = df["pct_str"].map({p: i for i, p in enumerate(PCT_ORDER)})
    return df.sort_values(["method", "_mord", "_pord"]).drop(columns=["_mord", "_pord"]).reset_index(drop=True)


def _overwrite_sheet(wb: openpyxl.Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    """Replace an existing sheet in wb with the contents of df."""
    if sheet_name in wb.sheetnames:
        idx = wb.sheetnames.index(sheet_name)
        del wb[sheet_name]
        ws = wb.create_sheet(sheet_name, idx)
    else:
        ws = wb.create_sheet(sheet_name)

    cols = list(df.columns)
    ws.append(cols)
    _style_header_row(ws)
    ws.freeze_panes = "A2"
    for i, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = ALT_ROW_FILL
    for col_idx, col_name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    print(f"    Rebuilt {sheet_name}: {len(df):,} rows")


def _rebuild_summary_sheets(wb: openpyxl.Workbook, df_iter: pd.DataFrame) -> None:
    """
    Recompute and overwrite Coef_Stability_Summary, Mean_Stability_*, and
    Model_Comparison sheets from the complete IterationDetail data.
    Used when the source workbook was built from a partial (resumed) run.
    """
    print("  Rebuilding summary sheets from IterationDetail...")
    df_css = _recompute_coef_stability_summary(df_iter)
    _overwrite_sheet(wb, "Coef_Stability_Summary", df_css)

    for mech in ["MCAR", "MAR", "NMAR"]:
        df_ms = _recompute_mean_stability(df_css, mech)
        _overwrite_sheet(wb, f"Mean_Stability_{mech}", df_ms)

    df_mc = _recompute_model_comparison(df_iter)
    _overwrite_sheet(wb, "Model_Comparison", df_mc)


# ---------------------------------------------------------------------------
# PDF generation — Paper_Info_Record.pdf  (Appendix A template)
# ---------------------------------------------------------------------------

# Maps paper_info xlsx Field names to Appendix A sections (order matters)
_PDF_SECTIONS = {
    "General Information": [
        "Paper ID", "Title", "Authors", "Year & Journal",
    ],
    "Data Structure": [
        "Data type", "Observations (N)", "Preferred model column",
    ],
    "Core Regression Model": [
        "Dependent variable", "Focal IV", "Controls",
        "Fixed effects", "Clustered SEs", "Weights",
    ],
    "Simulation Configuration": [
        "Key variables (3)", "Key variables (4)", "Key variables (5)",
        "MAR control", "Mechanisms", "Missingness proportions",
        "Methods", "Iterations per scenario", "MAR/NMAR strength",
        "M (multiple imputation)",
    ],
    "Baseline Validation": [
        "Baseline coefficient", "Baseline SE", "Baseline p-value",
        "Published value", "R-squared",
    ],
    "Run Log": [
        "Full run start", "Full run complete",
        "Total combinations", "Errors", "DL method", "MILGBM method", "Note",
    ],
}


_FIELD_ALIASES = {
    "Journal":       "Year & Journal",
    "Full run end":  "Full run complete",
    "Total runs":    "Total combinations",
    "QC status":     "Note",
}


def generate_paper_info_pdf(paper_id: str, cfg: dict) -> None:
    """
    Generate Paper_Info_Record.pdf matching RA_MISSING_DATA.pdf Appendix A.
    Saves to paper_dir/ and a disambiguated copy to repo root.
    """
    if not _REPORTLAB_OK:
        print("  SKIP PDF: reportlab not installed (pip install reportlab)")
        return

    author_year = cfg["author_year"]
    paper_dir   = cfg["paper_dir"]
    info_path   = cfg["paper_info"]

    # Load field→value dict from paper_info xlsx; normalise variant field names
    info_rows = _load_paper_info_rows(info_path)
    info = {_FIELD_ALIASES.get(field, field): value for field, value in info_rows}

    # Output paths
    pdf_local = paper_dir / "Paper_Info_Record.pdf"
    pdf_root  = REPO_ROOT / f"Paper_Info_Record_{author_year}.pdf"

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PaperTitle",
        parent=styles["Heading1"],
        fontSize=14, textColor=colors.HexColor("#1E50A0"),
        spaceAfter=4,
    )
    section_style = ParagraphStyle(
        "Section",
        parent=styles["Heading2"],
        fontSize=11, textColor=colors.HexColor("#1E50A0"),
        spaceBefore=10, spaceAfter=4,
        borderPad=2,
    )
    normal = styles["Normal"]
    normal.fontSize = 9

    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.grey)
        txt = f"Oxford Missing Data Study  |  {author_year}  |  Generated: {date.today()}"
        canvas.drawCentredString(A4[0] / 2, 1.5 * cm, txt)
        canvas.drawRightString(A4[0] - 2 * cm, 1.5 * cm, f"Page {doc.page}")
        canvas.restoreState()

    story = []

    # Title
    story.append(Paragraph(
        f"Paper Information Record — {author_year}", title_style
    ))
    story.append(Paragraph(
        f"Paper ID: {paper_id}  |  Oxford Missing Data Study", normal
    ))
    story.append(Spacer(1, 0.3 * cm))

    tbl_style = TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#1E50A0")),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 8),
        ("BACKGROUND",  (0, 1), (0, -1), colors.HexColor("#F0F0F8")),
        ("FONTNAME",    (0, 1), (0, -1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5FF")]),
        ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("VALIGN",      (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
        ("WORDWRAP",    (1, 0), (1, -1), True),
    ])

    for section_name, fields in _PDF_SECTIONS.items():
        # Filter to fields that exist in the info dict (skip empty Key variables variants)
        rows = [(f, info[f]) for f in fields if f in info and info.get(f, "")]
        if not rows:
            continue
        story.append(Paragraph(section_name, section_style))
        # Wrap values in Paragraph so long strings word-wrap inside the cell
        table_data = [["Field", "Value"]] + [
            [Paragraph(f, normal), Paragraph(str(v), normal)] for f, v in rows
        ]
        col_widths = [4.5 * cm, 12.5 * cm]
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(tbl_style)
        story.append(t)
        story.append(Spacer(1, 0.2 * cm))

    # Build once (flowables are stateful; building twice produces an empty PDF)
    doc = SimpleDocTemplate(
        str(pdf_local),
        pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2.5 * cm,
    )
    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    shutil.copy2(pdf_local, pdf_root)

    print(f"  PDF: Paper_Info_Record.pdf -> {pdf_local.parent.name}/ + root copy")


# ---------------------------------------------------------------------------
# Main per-paper builder
# ---------------------------------------------------------------------------
def process_paper(paper_id: str, cfg: dict) -> Path:
    print(f"\n=== Paper {paper_id} ({cfg['author_year']}) ===")

    source_wb_path = cfg["workbook"]
    if not source_wb_path.exists():
        raise FileNotFoundError(f"Source workbook not found: {source_wb_path}")

    print(f"  Loading source workbook: {source_wb_path.name}")
    wb = openpyxl.load_workbook(source_wb_path)
    print(f"  Sheets in source: {wb.sheetnames}")

    # Fix A — insert 00_PaperInfo as first sheet
    print("  Building 00_PaperInfo sheet...")
    if not cfg["paper_info"].exists():
        raise FileNotFoundError(f"paper_info not found: {cfg['paper_info']}")
    pir_rows = _load_paper_info_rows(cfg["paper_info"])
    _build_paper_info_sheet(wb, pir_rows)
    print(f"    00_PaperInfo: {len(pir_rows)} rows")

    # Fix B — build IterationDetail
    print("  Building IterationDetail sheet...")
    if cfg["txt_has_data"]:
        df_iter = _build_iteration_detail_0017(cfg)
    else:
        df_iter = _build_iteration_detail_0005(cfg)
    _write_iteration_detail_sheet(wb, df_iter)

    # Fix C — rebuild summary sheets if source workbook is from a resumed run
    if cfg.get("rebuild_summary"):
        _rebuild_summary_sheets(wb, df_iter)

    # Save to full_run/  — filename: {AuthorYear}_Report.xlsx
    out_name = f"{cfg['author_year']}_Report.xlsx"
    full_run_dir = cfg["paper_dir"] / "full_run"
    full_run_dir.mkdir(parents=True, exist_ok=True)
    out_path = full_run_dir / out_name
    print(f"  Saving to {out_path} ...")
    wb.save(str(out_path))

    size_mb = out_path.stat().st_size / 1_048_576
    print(f"  Saved: {out_path.name} ({size_mb:.1f} MB)")

    # Copy to repo root
    root_copy = REPO_ROOT / out_name
    shutil.copy2(out_path, root_copy)
    print(f"  Root copy: {root_copy.name} ({root_copy.stat().st_size / 1_048_576:.1f} MB)")

    # Generate Paper_Info_Record.pdf
    print("  Generating Paper_Info_Record.pdf...")
    generate_paper_info_pdf(paper_id, cfg)

    return root_copy


# ---------------------------------------------------------------------------
# Verification
# ---------------------------------------------------------------------------
def verify_output(paper_id: str, cfg: dict) -> None:
    author_year = cfg["author_year"]
    out_name     = f"{author_year}_Report.xlsx"
    root_path    = REPO_ROOT / out_name
    full_run_path = cfg["paper_dir"] / "full_run" / out_name
    pdf_local    = cfg["paper_dir"] / "Paper_Info_Record.pdf"
    pdf_root     = REPO_ROOT / f"Paper_Info_Record_{author_year}.pdf"

    ok = True
    for p in [root_path, full_run_path]:
        if not p.exists():
            print(f"  FAIL: {p} not found")
            ok = False
        elif p.stat().st_size < 1 * 1_048_576:
            print(f"  WARN: {p.name} is only {p.stat().st_size / 1_048_576:.1f} MB (expected >1 MB)")
        else:
            print(f"  OK:   {p.name} — {p.stat().st_size / 1_048_576:.1f} MB")

    for p in [pdf_local, pdf_root]:
        if p.exists():
            print(f"  OK:   {p.name}")
        else:
            print(f"  WARN: {p.name} not found (reportlab required)")

    if root_path.exists():
        wb = openpyxl.load_workbook(root_path, read_only=True)
        sheets = wb.sheetnames
        if sheets[0] != "00_PaperInfo":
            print(f"  FAIL: First sheet is '{sheets[0]}', expected '00_PaperInfo'")
            ok = False
        else:
            print(f"  OK:   First sheet = '00_PaperInfo'")

        if "IterationDetail" not in sheets:
            print(f"  FAIL: 'IterationDetail' sheet missing")
            ok = False
        else:
            ws_iter = wb["IterationDetail"]
            n_rows = ws_iter.max_row - 1
            print(f"  OK:   IterationDetail — {n_rows:,} rows")

        for sheet in ["Mean_Stability_MCAR", "Mean_Stability_MAR", "Mean_Stability_NMAR",
                      "Coef_Stability_Summary"]:
            if sheet in sheets:
                print(f"  OK:   {sheet} present")
            else:
                print(f"  WARN: {sheet} not found in output workbook")

        wb.close()

    if ok:
        print(f"  Verification PASSED for paper {paper_id} ({author_year})")
    else:
        print(f"  Verification had issues for paper {paper_id} ({author_year})")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="Generate AuthorYearReport deliverables")
    parser.add_argument("--paper", choices=list(PAPERS.keys()), default=None,
                        help="Process only this paper (default: all)")
    parser.add_argument("--pdf-only", action="store_true",
                        help="Regenerate only Paper_Info_Record PDFs (skip workbook processing)")
    args = parser.parse_args()

    papers_to_run = {args.paper: PAPERS[args.paper]} if args.paper else PAPERS

    if args.pdf_only:
        for paper_id, cfg in papers_to_run.items():
            print(f"\n=== Paper {paper_id} ({cfg['author_year']}) — PDF only ===")
            generate_paper_info_pdf(paper_id, cfg)
    else:
        for paper_id, cfg in papers_to_run.items():
            process_paper(paper_id, cfg)

        print("\n--- Verification ---")
        for paper_id, cfg in papers_to_run.items():
            verify_output(paper_id, cfg)

    print("\nDone.")


if __name__ == "__main__":
    main()
